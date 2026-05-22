"""
excel_loader.py — Ingesta de datos reales desde archivos Excel hacia el esquema legacy.

Propósito:
    Permite cargar datos reales y sucios directamente desde un archivo .xlsx a las
    tablas legacy, complementando el volumen sintético generado por Faker.
    Los datos se insertan *sin transformación* para que el pipeline de normalización
    (normalizer.py) los procese junto con el resto del conjunto de datos.

Mapeo de hojas → tablas legacy:
    Hoja "Biblioteca_Data"  → tabla Biblioteca_Data
    Hoja "Prestamos_Crudos" → tabla Prestamos_Crudos
    Hoja "Inventario_Sedes" → tabla Inventario_Sedes
    Hoja "Reseñas_Usuarios" → tabla Reseñas_Usuarios

Cualquier hoja cuyo nombre no coincida con una entrada del mapeo se omite con un aviso,
garantizando robustez ante archivos con hojas adicionales de auditoría o metadatos.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Mapeo: nombre de hoja → (nombre de tabla SQL, columnas esperadas en orden)
# Las columnas deben coincidir con la primera fila del encabezado del Excel.
# ---------------------------------------------------------------------------
_SHEET_MAP: dict[str, tuple[str, list[str]]] = {
    "Biblioteca_Data": (
        '"Biblioteca_Data"',
        ["titulo_libro", "autor_nombre", "categoria_y_descripcion", "editorial_info", "fecha_publicacion"],
    ),
    "Prestamos_Crudos": (
        '"Prestamos_Crudos"',
        ["id_prestamo", "nombre_usuario", "correo_usuario", "libros_prestados", "fecha_salida", "estado_prestamo"],
    ),
    "Inventario_Sedes": (
        '"Inventario_Sedes"',
        ["sede_nombre", "ubicacion_sede", "libro_asociado", "cantidad_total"],
    ),
    "Reseñas_Usuarios": (
        '"Reseñas_Usuarios"',
        ["usuario_id", "libro_titulo", "comentario", "calificacion"],
    ),
}


def _build_insert(table: str, columns: list[str]) -> str:
    """Construye dinámicamente el INSERT parametrizado para psycopg2."""
    cols = ", ".join(columns)
    placeholders = ", ".join(["%s"] * len(columns))
    return f"INSERT INTO {table} ({cols}) VALUES ({placeholders})"


def _column_indexes(header: tuple[Any, ...] | None, expected_cols: list[str]) -> list[int]:
    """Obtiene las posiciones de las columnas esperadas usando el encabezado."""
    header_map = {
        str(column).strip(): index
        for index, column in enumerate(header or ())
        if column is not None
    }
    if all(column in header_map for column in expected_cols):
        return [header_map[column] for column in expected_cols]

    logger.warning(
        "Encabezado incompleto o inesperado (%s). Se usará mapeo posicional.",
        header,
    )
    return list(range(len(expected_cols)))


def load_excel(conn, xlsx_path: str | Path) -> dict[str, int]:
    """
    Carga cada hoja del archivo Excel en la tabla legacy correspondiente.

    Los valores se insertan tal como vienen del archivo (raw), incluyendo
    datos nulos, formatos incorrectos y combinaciones inválidas, ya que la
    responsabilidad de limpieza recae íntegramente en el pipeline ETL.

    Args:
        conn:      Conexión psycopg2 activa (manejada por el context manager de db.py).
        xlsx_path: Ruta absoluta o relativa al archivo .xlsx a procesar.

    Returns:
        Diccionario {nombre_hoja: filas_insertadas} para integración en el
        reporte de estadísticas del orquestador principal.

    Raises:
        FileNotFoundError: Si el archivo .xlsx no existe en la ruta indicada.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo Excel no encontrado: {path}")

    stats: dict[str, int] = {}

    # Carga en modo solo-valores (data_only=True) para evitar leer fórmulas
    # crudas en lugar de los valores calculados almacenados.
    workbook = openpyxl.load_workbook(path, data_only=True)

    with conn.cursor() as cur:
        for sheet_name in workbook.sheetnames:
            if sheet_name not in _SHEET_MAP:
                logger.warning(
                    "Hoja '%s' no tiene mapeo definido — se omite.", sheet_name
                )
                stats[sheet_name] = 0
                continue

            table_name, expected_cols = _SHEET_MAP[sheet_name]
            sheet = workbook[sheet_name]
            sql = _build_insert(table_name, expected_cols)

            rows_inserted = 0
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            indexes = _column_indexes(header, expected_cols)

            for row in rows:
                # Omitir filas completamente vacías (filas en blanco del Excel)
                if all(cell is None for cell in row):
                    continue

                # Tomar valores por nombre de columna. Esto evita desplazamientos
                # cuando el Excel trae columnas extra, como id_registro.
                values: list[Any] = [
                    row[index] if index < len(row) else None
                    for index in indexes
                ]

                try:
                    cur.execute("SAVEPOINT load_row")
                    cur.execute(sql, values)
                    cur.execute("RELEASE SAVEPOINT load_row")
                    rows_inserted += 1
                except Exception as exc:
                    cur.execute("ROLLBACK TO SAVEPOINT load_row")
                    # Registrar la fila problemática sin abortar la carga completa.
                    logger.warning(
                        "Fila omitida en hoja '%s' (error: %s) — valores: %s",
                        sheet_name, exc, values,
                    )

            stats[sheet_name] = rows_inserted
            logger.info(
                "Excel → %s: %d filas insertadas.", sheet_name, rows_inserted
            )

    workbook.close()
    return stats
