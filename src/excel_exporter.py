"""excel_exporter.py — Exporta datos normalizados a Excel.

Este módulo se conecta a la base de datos PostgreSQL (utilizando la conexión
pasada por `pg_connection`) y exporta las seis tablas normalizadas a un
archivo .xlsx. Cada tabla tiene su propia hoja, los encabezados están en
negrita y las columnas se auto‑ajustan al contenido.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

# Nombre de las tablas normalizadas en el esquema 3FN. Debe coincidir con
# los nombres usados en los scripts SQL de normalización.
NORMALIZED_TABLES = [
    "libros",
    "autores",
    "usuarios",
    "prestamos",
    "inventario",
    "resenas",
]


def _fetch_table(conn, table: str) -> Sequence[Sequence]:
    """Devuelve todas las filas de *table* como una lista de tuplas.

    Se asume que la tabla ya existe y que el cursor de *conn* soporta
    ``fetchall``. El orden de columnas es el que devuelve PostgreSQL por
    defecto (es decir, el orden del esquema)."""
    with conn.cursor() as cur:
        cur.execute(f"SELECT * FROM {table}")
        rows = cur.fetchall()
        # Obtener nombres de columnas a partir de la descripción del cursor
        col_names = [desc[0] for desc in cur.description]
    return col_names, rows


def _auto_adjust_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Ajusta el ancho de cada columna al contenido máximo.

    openpyxl no tiene una función nativa; calculamos la longitud máxima de
    cada celda (incluyendo el encabezado) y asignamos esa longitud al ancho.
    """
    for column_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = length + 2  # margen extra


def export_normalized_to_excel(conn, output_path: Path | str) -> Path:
    """Exporta las seis tablas normalizadas a *output_path*.

    Parámetros
    ----------
    conn: objeto de conexión ``psycopg2`` abierto.
    output_path: ruta donde se guardará el archivo .xlsx.

    Retorna
    -------
    Path al archivo generado.
    """
    output_path = Path(output_path)
    workbook = openpyxl.Workbook()
    # openpyxl crea una hoja por defecto; la eliminamos porque la vamos a
    # crear explícitamente para cada tabla.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table in NORMALIZED_TABLES:
        logger.info("Exportando tabla %s a hoja Excel", table)
        headers, rows = _fetch_table(conn, table)
        ws = workbook.create_sheet(title=table)
        # Escribir encabezados en negrita
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
        # Escribir filas
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        # Auto‑ajustar ancho de columnas
        _auto_adjust_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    logger.info("Archivo Excel exportado: %s", output_path)
    return output_path
